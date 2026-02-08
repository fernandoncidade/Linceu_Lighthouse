import os
from utils.LogManager import LogManager
logger = LogManager.get_logger()

def extrair_metadados_planilha(caminho, loc):
    metadados = {}
    ext = os.path.splitext(caminho)[1].lower()

    try:
        if ext in ['.xlsx', '.xlsm', '.xls']:
            try:
                if ext in ['.xlsx', '.xlsm']:
                    import zipfile
                    if not zipfile.is_zipfile(caminho):
                        logger.error(f"Arquivo XLSX inválido (não é ZIP): {caminho}")
                        return metadados

                    from openpyxl import load_workbook
                    try:
                        wb = load_workbook(caminho, read_only=True, data_only=True)
                        planilhas = len(wb.sheetnames)
                        metadados['planilhas'] = planilhas
                        metadados['nomes_planilhas'] = ", ".join(wb.sheetnames)

                        linhas_total = 0
                        colunas_total = 0

                        for sheet_name in wb.sheetnames[:3]:
                            sheet = wb[sheet_name]
                            max_row = getattr(sheet, 'max_row', 0) or 0
                            max_col = getattr(sheet, 'max_column', 0) or 0
                            try:
                                linhas_total += int(max_row)

                            except Exception:
                                pass

                            try:
                                max_col_int = int(max_col)
                                if max_col_int > colunas_total:
                                    colunas_total = max_col_int

                            except Exception:
                                pass

                        metadados['total_linhas'] = str(linhas_total)
                        metadados['colunas'] = str(colunas_total)

                        is_protected = False
                        for sheet_name in wb.sheetnames:
                            sheet = wb[sheet_name]
                            if hasattr(sheet, 'protection') and sheet.protection.sheet:
                                is_protected = True
                                break

                        if is_protected:
                            metadados['protegido'] = loc.get_text("yes") + " (planilhas protegidas)"

                        if wb.properties:
                            if wb.properties.creator:
                                metadados['autor'] = wb.properties.creator

                            if wb.properties.title:
                                metadados['titulo'] = wb.properties.title

                            if wb.properties.created:
                                metadados['data_criacao_doc'] = str(wb.properties.created)

                            if wb.properties.modified:
                                metadados['data_mod_doc'] = str(wb.properties.modified)

                        wb.close()

                    except Exception as e_xlsx:
                        try:
                            import xml.etree.ElementTree as ET
                            with zipfile.ZipFile(caminho, 'r') as z:
                                if 'xl/workbook.xml' in z.namelist():
                                    xml_bytes = z.read('xl/workbook.xml')
                                    root = ET.fromstring(xml_bytes)
                                    ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                                    sheets = root.findall('.//ns:sheets/ns:sheet', ns) or root.findall('.//sheets/sheet')
                                    metadados['planilhas'] = len(sheets)
                                    metadados['nomes_planilhas'] = ", ".join([s.get('name', '') for s in sheets if s is not None and s.get('name')])

                        except Exception as e_xml:
                            logger.error(f"Erro no fallback XML para XLSX {caminho}: {e_xml}", exc_info=True)

                elif ext == '.xls':
                    import xlrd
                    try:
                        wb = xlrd.open_workbook(caminho, on_demand=True)
                        planilhas = len(wb.sheet_names())
                        metadados['planilhas'] = planilhas
                        metadados['nomes_planilhas'] = ", ".join(wb.sheet_names())

                        linhas_total = 0
                        colunas_total = 0

                        for idx in range(min(3, planilhas)):
                            sheet = wb.sheet_by_index(idx)
                            linhas_total += sheet.nrows
                            if sheet.ncols > colunas_total:
                                colunas_total = sheet.ncols

                        metadados['total_linhas'] = str(linhas_total)
                        metadados['colunas'] = str(colunas_total)

                        if hasattr(wb, 'protection_mode') and wb.protection_mode:
                            metadados['protegido'] = loc.get_text("yes")

                        wb.release_resources()

                    except Exception as e_xls:
                        msg = str(e_xls)
                        if 'Workbook is encrypted' in msg:
                            metadados['protegido'] = loc.get_text("yes") + " (senha)"

                        elif 'Excel xlsx file; not supported' in msg:
                            import zipfile, xml.etree.ElementTree as ET
                            if zipfile.is_zipfile(caminho):
                                try:
                                    with zipfile.ZipFile(caminho, 'r') as z:
                                        if 'xl/workbook.xml' in z.namelist():
                                            xml_bytes = z.read('xl/workbook.xml')
                                            root = ET.fromstring(xml_bytes)
                                            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                                            sheets = root.findall('.//ns:sheets/ns:sheet', ns) or root.findall('.//sheets/sheet')
                                            metadados['planilhas'] = len(sheets)
                                            metadados['nomes_planilhas'] = ", ".join([s.get('name', '') for s in sheets if s is not None and s.get('name')])

                                except Exception as e_xml2:
                                    logger.error(f"Erro no fallback XML para XLS renomeado {caminho}: {e_xml2}", exc_info=True)

                        else:
                            logger.error(f"Erro ao extrair metadados da planilha {caminho}: {e_xls}", exc_info=True)

            except Exception as e:
                logger.error(f"Erro ao extrair metadados da planilha {caminho}: {e}", exc_info=True)

        elif ext == '.csv':
            try:
                import csv
                linhas = 0
                colunas = 0

                with open(caminho, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    primeira_linha = next(reader, None)
                    if primeira_linha:
                        colunas = len(primeira_linha)
                        linhas = 1

                    for _ in reader:
                        linhas += 1

                metadados['planilhas'] = 1
                nome_arquivo = os.path.basename(caminho)
                nome_sem_ext = os.path.splitext(nome_arquivo)[0]
                metadados['nomes_planilhas'] = nome_sem_ext

                metadados['total_linhas'] = linhas
                metadados['colunas'] = colunas

            except Exception as e:
                logger.error(f"Erro ao extrair metadados do CSV {caminho}: {e}", exc_info=True)

    except Exception as e:
        logger.error(f"Erro geral ao extrair metadados da planilha {caminho}: {e}", exc_info=True)

    return metadados
