import os
import win32file
import win32con
import ctypes
from ctypes import wintypes
from source.utils.LogManager import LogManager
logger = LogManager.get_logger()

def _arquivo_assinado(caminho: str) -> bool:
    try:
        TRUST_E_NOSIGNATURE = 0x800B0100
        WTD_UI_NONE = 2
        WTD_REVOKE_NONE = 0
        WTD_CHOICE_FILE = 1
        WTD_STATEACTION_IGNORE = 0

        class GUID(ctypes.Structure):
            _fields_ = [
                ("Data1", wintypes.DWORD),
                ("Data2", wintypes.WORD),
                ("Data3", wintypes.WORD),
                ("Data4", ctypes.c_ubyte * 8),
            ]

            def __init__(self, d1, d2, d3, d4):
                super().__init__(d1, d2, d3, (ctypes.c_ubyte * 8)(*d4))

        class WINTRUST_FILE_INFO(ctypes.Structure):
            _fields_ = [
                ("cbStruct", wintypes.DWORD),
                ("pcwszFilePath", wintypes.LPCWSTR),
                ("hFile", wintypes.HANDLE),
                ("pgKnownSubject", ctypes.POINTER(GUID)),
            ]

        class WINTRUST_DATA_UNION(ctypes.Union):
            _fields_ = [("pFile", ctypes.POINTER(WINTRUST_FILE_INFO))]

        class WINTRUST_DATA(ctypes.Structure):
            _anonymous_ = ("u",)
            _fields_ = [
                ("cbStruct", wintypes.DWORD),
                ("pPolicyCallbackData", ctypes.c_void_p),
                ("pSIPClientData", ctypes.c_void_p),
                ("dwUIChoice", wintypes.DWORD),
                ("fdwRevocationChecks", wintypes.DWORD),
                ("dwUnionChoice", wintypes.DWORD),
                ("u", WINTRUST_DATA_UNION),
                ("dwStateAction", wintypes.DWORD),
                ("hWVTStateData", wintypes.HANDLE),
                ("pwszURLReference", wintypes.LPCWSTR),
                ("dwProvFlags", wintypes.DWORD),
                ("dwUIContext", wintypes.DWORD),
            ]

        action_guid = GUID(0x00AAC56B, 0xCD44, 0x11D0, (0x8C, 0xC2, 0x00, 0xC0, 0x4F, 0xC2, 0x95, 0xEE))

        wffi = WINTRUST_FILE_INFO()
        wffi.cbStruct = ctypes.sizeof(WINTRUST_FILE_INFO)
        wffi.pcwszFilePath = caminho
        wffi.hFile = None
        wffi.pgKnownSubject = None

        wtd = WINTRUST_DATA()
        wtd.cbStruct = ctypes.sizeof(WINTRUST_DATA)
        wtd.pPolicyCallbackData = None
        wtd.pSIPClientData = None
        wtd.dwUIChoice = WTD_UI_NONE
        wtd.fdwRevocationChecks = WTD_REVOKE_NONE
        wtd.dwUnionChoice = WTD_CHOICE_FILE
        wtd.pFile = ctypes.pointer(wffi)
        wtd.dwStateAction = WTD_STATEACTION_IGNORE
        wtd.hWVTStateData = None
        wtd.pwszURLReference = None
        wtd.dwProvFlags = 0
        wtd.dwUIContext = 0

        WinVerifyTrust = ctypes.windll.wintrust.WinVerifyTrust
        WinVerifyTrust.argtypes = [wintypes.HWND, ctypes.POINTER(GUID), ctypes.POINTER(WINTRUST_DATA)]
        WinVerifyTrust.restype = wintypes.LONG

        result = WinVerifyTrust(None, ctypes.byref(action_guid), ctypes.byref(wtd))
        if result == 0:
            return True

        return False

    except Exception as e:
        logger.error(f"Erro ao verificar assinatura do arquivo {caminho}: {e}", exc_info=True)
        return False

def get_protecao_arquivo(gerenciador, item, loc):
    caminho = item.get("dir_atual") or item.get("dir_anterior")
    if not caminho or not os.path.exists(caminho):
        if "protegido" in item and item["protegido"]:
            return loc.traduzir_metadados(item["protegido"], "protegido")

        return ""

    try:
        with gerenciador.lock_cache:
            if (
                hasattr(gerenciador, "cache_metadados")
                and caminho in gerenciador.cache_metadados
                and "protegido" in gerenciador.cache_metadados[caminho]
            ):
                protegido_cache = gerenciador.cache_metadados[caminho]["protegido"]
                return loc.get_text("yes") if bool(protegido_cache) else loc.get_text("no")

        attrs = win32file.GetFileAttributes(caminho)
        is_readonly = bool(attrs & win32con.FILE_ATTRIBUTE_READONLY)
        is_hidden = bool(attrs & win32con.FILE_ATTRIBUTE_HIDDEN)
        is_system = bool(attrs & win32con.FILE_ATTRIBUTE_SYSTEM)
        is_encrypted = bool(attrs & win32con.FILE_ATTRIBUTE_ENCRYPTED)
        is_compressed = bool(attrs & win32con.FILE_ATTRIBUTE_COMPRESSED)

        protegido = any([is_readonly, is_hidden, is_system, is_encrypted, is_compressed])

        if os.path.isfile(caminho):
            if not protegido and _arquivo_assinado(caminho):
                protegido = True

        if not protegido and _arquivo_assinado(caminho):
            protegido = True

        if not protegido:
            ext = os.path.splitext(caminho)[1].lower()
            if ext in ['.docx', '.xlsx', '.pptx', '.doc', '.xls', '.ppt']:
                try:
                    if ext == '.docx':
                        try:
                            from docx import Document
                            Document(caminho)

                        except Exception as e:
                            if "document is encrypted" in str(e).lower():
                                protegido = True

                    elif ext == '.xlsx':
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook(caminho, read_only=True, data_only=True)
                            for sheet_name in wb.sheetnames:
                                sheet = wb[sheet_name]
                                if hasattr(sheet, 'protection') and getattr(sheet.protection, 'sheet', False):
                                    protegido = True
                                    break

                            wb.close()

                        except Exception as e:
                            if "file is encrypted" in str(e).lower():
                                protegido = True

                    elif ext == '.pptx':
                        try:
                            from pptx import Presentation
                            Presentation(caminho)

                        except Exception as e:
                            if "is encrypted" in str(e).lower():
                                protegido = True

                    elif ext in ['.doc', '.xls', '.ppt']:
                        try:
                            import olefile
                            if olefile.isOleFile(caminho):
                                try:
                                    with olefile.OleFileIO(caminho):
                                        pass

                                except Exception as e:
                                    if "encrypted" in str(e).lower():
                                        protegido = True

                        except ImportError:
                            pass

                except ImportError:
                    pass

            if not protegido and ext in ['.zip', '.rar', '.7z']:
                try:
                    if ext == '.zip':
                        try:
                            import zipfile
                            with zipfile.ZipFile(caminho, 'r') as zf:
                                for info in zf.infolist():
                                    if getattr(info, "flag_bits", 0) & 0x1:
                                        protegido = True
                                        break

                        except Exception as e:
                            if any(s in str(e).lower() for s in ["password", "encrypted", "decryption"]):
                                protegido = True

                    elif ext == '.rar':
                        try:
                            import rarfile
                            with rarfile.RarFile(caminho) as rf:
                                try:
                                    if hasattr(rf, "needs_password") and rf.needs_password():
                                        protegido = True

                                    else:
                                        for info in rf.infolist():
                                            needs_pwd = getattr(info, "needs_password", None)
                                            if callable(needs_pwd) and needs_pwd():
                                                protegido = True
                                                break

                                except Exception as e:
                                    if any(s in str(e).lower() for s in ["password", "encrypted", "decryption"]):
                                        protegido = True

                        except ImportError:
                            pass

                    elif ext == '.7z':
                        try:
                            import py7zr
                            from py7zr import exceptions as py7ex
                            with py7zr.SevenZipFile(caminho, mode='r') as z:
                                try:
                                    needs_pwd = getattr(z, "needs_password", None)
                                    if callable(needs_pwd) and needs_pwd():
                                        protegido = True

                                    else:
                                        infos = z.list()
                                        for info in infos:
                                            if getattr(info, "encrypted", False) or getattr(info, "is_encrypted", False):
                                                protegido = True
                                                break

                                except Exception as e:
                                    if isinstance(e, getattr(py7ex, "PasswordRequired", tuple())):
                                        protegido = True

                                    elif any(s in str(e).lower() for s in ["password", "encrypted", "decryption"]):
                                        protegido = True

                        except ImportError:
                            pass

                except Exception as e:
                    logger.error(f"Erro ao verificar proteção de arquivo compactado {caminho}: {e}", exc_info=True)

        try:
            with gerenciador.lock_cache:
                if hasattr(gerenciador, "cache_metadados"):
                    entry = gerenciador.cache_metadados.setdefault(caminho, {})
                    entry["protegido"] = bool(protegido)

        except Exception as e:
            logger.error(f"Erro ao salvar proteção em cache: {e}", exc_info=True)

        return loc.get_text("yes") if protegido else loc.get_text("no")

    except Exception as e:
        logger.error(f"Erro ao verificar proteção do arquivo {caminho}: {e}", exc_info=True)
        if "protegido" in item and item["protegido"]:
            return loc.traduzir_metadados(item["protegido"], "protegido")

        print(f"Erro ao verificar proteção: {e}")
        return ""
